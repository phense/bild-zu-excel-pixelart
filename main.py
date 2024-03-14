from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import sys
import os

def pixelate_image(image_path, pixel_size):
    img = Image.open(image_path)

    # Calculate the size of the pixelated version
    small_size = (img.width // pixel_size, img.height // pixel_size)

    # Resize the image to small size and back to original size
    img_small = img.resize(small_size, Image.NEAREST)
    img_pixelated = img_small.resize(img.size, Image.NEAREST)

    # Convert to RGB format if not already
    img_pixelated = img_pixelated.convert('RGB')

    # Extract color values
    pixel_values = []
    for y in range(0, img_pixelated.height, pixel_size):
        row = []
        for x in range(0, img_pixelated.width, pixel_size):
            r, g, b = img_pixelated.getpixel((x, y))
            row.append((r, g, b))
        pixel_values.append(row)

    return pixel_values

def export_to_excel(color_grid, filename='pixelated_image.xlsx'):
    wb = Workbook()
    sheet = wb.active
    column_width = 2.75
    
    # Spaltenbreite setzen
    num_columns = len(color_grid[0])
    for col_idx in range(1, num_columns + 1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = column_width
        
    # Farben in Excel eintragen
    for row_idx, row in enumerate(color_grid, start=1):
        for col_idx, color in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            # Convert RGB to hex
            hex_color = f"{color[0]:02x}{color[1]:02x}{color[2]:02x}"
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    wb.save(filename)

def main(folder_path): 
    folder_path = folder_path.replace('\\', '/')
    pixel_size = 10 
    
    print(f"Verarbeiteter Ordner: {folder_path}")  
    # Read all image files from the specified folder    
    image_files = [file for file in os.listdir(folder_path) 
                   if file.lower().endswith(('.jpg', '.jpeg', '.png'))]
    
    if not image_files:
        print("No image files found in the specified folder.")
        return
    
    # Create the Excel workbook
    wb = Workbook() 
    # Remove the default sheet created by openpyxl
    #wb.remove(wb.active)  
    
    for image_file in image_files:
        image_path = os.path.join(folder_path, image_file)
        filename = image_path + '.xlsx'
        color_grid = pixelate_image(image_path, pixel_size)
        # Use the filename without its extension as the sheet name
        sheet_name = os.path.splitext(image_file)[0]
        sheet = wb.create_sheet(title=sheet_name[:31])
        # Sheet names in Excel are limited to 31 characters
        export_to_excel(color_grid, filename)
   
if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Bitte geben Sie den Pfad des Ordners als Parameter an.")
    else:
        folder_path = sys.argv[1]
        main(folder_path)

# The 'color_grid' variable now holds your grid of color values